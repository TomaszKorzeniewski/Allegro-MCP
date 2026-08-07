#!/usr/bin/env python3
"""Excel: unikalne SKU (sygnatury) ofert POJEDYNCZYCH z konta Allegro + stan do weryfikacji."""
import sys, re
from _sciezki import KORZEN, WYNIKI  # ustawia sys.path na korzeń projektu
from allegro_client import AllegroClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

c = AllegroClient()

def fetch(status):
    out = []; off = 0
    while True:
        r = c.get("/sale/offers", params={"limit": 1000, "offset": off, "publication.status": status})
        b = r.get("offers", []); out += b; tc = r.get("totalCount", 0); off += len(b)
        if off >= tc or not b:
            break
    return out

offers = fetch("ACTIVE") + fetch("INACTIVE")

def is_set(ext, name):
    e = ext or ""; n = name or ""
    if re.search(r"-\s*\d+\s*szt", e): return True
    if re.search(r"\bx\s*\d+\b", e): return True
    if re.match(r"^\s*(zestaw|\d+\s*szt|\d+\s*x|\d+pak|\d+\s*sztuk)", n, re.I): return True
    if re.search(r"zestaw", n, re.I): return True
    return False

def grupa(ext):
    if ext.startswith("KT"): return "Tuba tekturowa"
    if ext.startswith("Koszyk"): return "Koszyk na piwo"
    if ext.startswith(("TP/", "TPP", "TPU")): return "Taśma pakowa"
    return "Inne"

singles = {}
for o in offers:
    ext = (o.get("external") or {}).get("id")
    if not ext or is_set(ext, o.get("name", "")):
        continue
    singles.setdefault(ext, []).append({
        "oid": o.get("id"),
        "name": o.get("name", ""),
        "stock": (o.get("stock") or {}).get("available"),
        "pub": (o.get("publication") or {}).get("status"),
        "price": ((o.get("sellingMode") or {}).get("price") or {}).get("amount"),
    })

# sort: taśmy TP, potem TPP/TPU, potem KT, potem Koszyk
def sortkey(ext):
    g = {"Taśma pakowa": 0, "Tuba tekturowa": 1, "Koszyk na piwo": 2, "Inne": 3}[grupa(ext)]
    return (g, ext)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SKU pojedyncze"
headers = ["Lp.", "SKU (sygnatura)", "Grupa", "Nazwa oferty", "Cena (zł)",
           "Stan wg Allegro (szt.)", "Ilość zweryfikowana", "Uwagi"]
ws.append(headers)

hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
verify_fill = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for col, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

lp = 0
for ext in sorted(singles, key=sortkey):
    rows = sorted(singles[ext], key=lambda r: -(r["stock"] or 0))
    prim = rows[0]
    uwaga = ""
    if len(rows) > 1:
        extra = "; ".join(f"{r['oid']} (stan {r['stock']}, {r['pub']})" for r in rows[1:])
        uwaga = f"DUPLIKAT — druga oferta: {extra}"
    lp += 1
    ws.append([lp, ext, grupa(ext), prim["name"], prim["price"],
               prim["stock"], "", uwaga])
    r = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=r, column=col).border = border
    ws.cell(row=r, column=7).fill = verify_fill  # kolumna do wpisania

widths = [5, 22, 16, 52, 10, 20, 20, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ws.max_row}"

out = str(WYNIKI / "SKU_pojedyncze_do_weryfikacji.xlsx")
wb.save(out)
print(f"Zapisano: {out}")
print(f"Unikalnych SKU pojedynczych: {lp}")
